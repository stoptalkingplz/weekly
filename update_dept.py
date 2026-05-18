# -*- coding: utf-8 -*-
"""
逐层上报周报脚本（从叶子部门 atomic JSON 动态汇总，无 weekly_state 依赖）

设计边界：
1. 本脚本只负责非叶子/父部门的逐层上报周报。
2. 数据源始终来自目标父部门下所有叶子部门的每日 atomic JSON。
3. 输入前删除个人 mention，父部门周报不输出个人信息。
4. 支持大数据量分 batch：platform -> project batch -> 超大 project 按 leaf_dept 拆分。
5. 云端平台环境无需 main，直接执行本脚本即可。
"""

import builtins
import os
import re
import sys
import csv
import json
import time
import uuid
import tempfile
import traceback
from copy import deepcopy
from pathlib import Path
from datetime import datetime, timedelta
from collections import OrderedDict, defaultdict
from concurrent.futures import ThreadPoolExecutor, as_completed

import requests
from openai import OpenAI
from zenv import get_zdkit_env
from zdbase import ZFile  # 保留平台兼容；本脚本主体不直接依赖

# =============================================================================
# 云端运行环境兼容：print flush + UTF-8 日志
# =============================================================================
try:
    sys.stdout.reconfigure(encoding="utf-8")
    sys.stderr.reconfigure(encoding="utf-8")
except Exception:
    pass

if not getattr(builtins.print, "_patched_flush", False):
    _original_print = builtins.print

    def print(*args, **kwargs):
        kwargs.setdefault("flush", True)
        _original_print(*args, **kwargs)

    print._patched_flush = True
    builtins.print = print


def safe_log_text(text):
    text = str(text or "")
    try:
        text.encode("utf-8")
        return text
    except Exception:
        return text.encode("utf-8", errors="replace").decode("utf-8", errors="replace")


# =============================================================================
# 全局配置加载
# =============================================================================
zenv_obj = get_zdkit_env()
BASE_URL = zenv_obj.zdkit._http_client.config.get("url")

try:
    with open(config_file.path, "r", encoding="utf-8") as config_fp:
        config = json.load(config_fp)
except Exception as e:
    print(f"❌ 配置文件读取失败: {e}")
    raise

AK = config.get("ak")
SK = config.get("sk")
ORG_GUID = config.get("org_guid")
USER_GUID = config.get("user_guid")
generate_type = "rollup_weekly"

# =============================================================================
# 推荐 config 关键字段示例
# =============================================================================
# {
#   "llm_global": {
#     "base_url": "http://xxx/cloud/v1",
#     "api_key": "xxx",
#     "model": "doubao-seed-2.0-pro",
#     "temperature": 0.3,
#     "max_tokens": 4096,
#     "print_stream": false,
#     "max_retries": 5
#   },
#   "rollup_global": {
#     "rollup_trend_prompt_file_guid": "父部门趋势分析txt文件guid",
#     "rollup_final_prompt_file_guid": "父部门最终整理txt文件guid",
#     "rollup_validation_prompt_file_guid": "父部门覆盖校验txt文件guid",
#     "rollup_repair_prompt_file_guid": "父部门遗漏修复txt文件guid",
#     "rollup_projects_per_batch": 5,
#     "rollup_max_batch_chars": 12000,
#     "rollup_tree_max_chars": 2000,
#     "rollup_trend_temperature": 0.1,
#     "rollup_final_temperature": 0.2,
#     "rollup_validation_temperature": 0.0,
#     "enable_rollup_coverage_check": true,
#     "enable_rollup_second_validation": false,
#     "batch_number": 10
#   },
#   "rollup_target_dept_ids": ["CET"],
#   "dept_config_file_guid": "部门配置Excel文件guid"
# }
#
# Prompt GUID 也兼容放在 config 顶层：
# "rollup_trend_prompt_file_guid": "...",
# "rollup_final_prompt_file_guid": "...",
# "rollup_validation_prompt_file_guid": "...",
# "rollup_repair_prompt_file_guid": "..."



# =============================================================================
# 配置读取：全局默认 + 局部覆盖
# =============================================================================
def get_llm_config(key, default=None):
    llm_global = config.get("llm_global", {}) or {}
    if llm_global.get(key) is not None:
        return llm_global.get(key)

    legacy_key_map = {
        "base_url": "llm_base_url",
        "api_key": "llm_api_key",
        "model": "llm_model",
        "temperature": "llm_temperature",
        "max_tokens": "llm_max_tokens",
        "print_stream": "llm_print_stream",
        "max_retries": "llm_max_retries",
    }
    legacy_key = legacy_key_map.get(key)
    if legacy_key and config.get(legacy_key) is not None:
        return config.get(legacy_key)
    return default


def get_rollup_config(local_conf, key, default=None):
    if local_conf and local_conf.get(key) is not None:
        return local_conf.get(key)
    rollup_global = config.get("rollup_global", {}) or {}
    if rollup_global.get(key) is not None:
        return rollup_global.get(key)
    weekly_global = config.get("weekly_global", {}) or {}
    if weekly_global.get(key) is not None:
        return weekly_global.get(key)
    if config.get(key) is not None:
        return config.get(key)
    return default


def normalize_to_list(value):
    if value is None:
        return []
    if isinstance(value, list):
        return [x for x in value if x]
    if isinstance(value, str):
        return [value] if value.strip() else []
    return []


# =============================================================================
# OpenAI SDK 直连模型配置
# =============================================================================
LLM_BASE_URL = get_llm_config("base_url", "")
LLM_API_KEY = get_llm_config("api_key", "")
LLM_MODEL = get_llm_config("model", "doubao-seed-2.0-pro")
LLM_TEMPERATURE = float(get_llm_config("temperature", 0.3))
LLM_MAX_TOKENS = int(get_llm_config("max_tokens", 4096))
LLM_PRINT_STREAM = bool(get_llm_config("print_stream", False))
LLM_MAX_RETRIES = int(get_llm_config("max_retries", 5))

if not LLM_BASE_URL or not LLM_API_KEY:
    raise ValueError("请在 config.llm_global 中配置 base_url 和 api_key，或兼容旧字段 llm_base_url / llm_api_key")

openai_client = OpenAI(base_url=LLM_BASE_URL, api_key=LLM_API_KEY)


# =============================================================================
# API 路由
# =============================================================================
ACCESS_TOKEN_ROUTE = "/api/user/platform/getAccessToken"
NOTE_JSON_ROUTE = "/platform/ws/noteInfo/getDocJson"
DOC_TREE_ROUTE = "/platform/api/main/doc/treeList"
SIGNED_URL_ROUTE = "/platform/api/main/storage/getSignedUrl"
WORKSPACE_SAVE_ROUTE = "/middle/server/api/workspace/save"
MD_INSERT_ROUTE = "/middle/server/api/file/md/insert"
MESSAGE_SEND_ROUTE = "/middle/server/api/msg/send"

MESSAGE_TEMPLATE_ID = "80"
PLATFORM_TYPE = "all"


# =============================================================================
# 通用工具函数
# =============================================================================
def get_headers_with_ak(user_guid="", doc_id=""):
    response = requests.post(
        url=BASE_URL + ACCESS_TOKEN_ROUTE,
        json={"ak": AK, "sk": SK},
        timeout=30,
    )
    response_json = response.json()
    if not response_json.get("data"):
        raise Exception(f"获取 AccessToken 失败: {response_json}")

    access_token = response_json["data"].get("accessToken")
    headers = {
        "Access-Token": access_token,
        "ak": AK,
        "X-User-GUID": user_guid or USER_GUID,
    }
    if doc_id:
        headers["docId"] = doc_id
    return headers


def strip_markdown_wrapper(content):
    content = (content or "").strip()
    if content.startswith("```json"):
        content = content[len("```json"):].lstrip("\n")
    elif content.startswith("```markdown"):
        content = content[len("```markdown"):].lstrip("\n")
    elif content.startswith("```"):
        content = content[3:].lstrip("\n")
    if content.endswith("```"):
        content = content[:-3].rstrip("\n")
    return content.strip()


def safe_json_loads(text):
    clean_text = strip_markdown_wrapper(text)
    try:
        return json.loads(clean_text)
    except Exception:
        pass
    object_match = re.search(r"(\{.*\})", clean_text, flags=re.DOTALL)
    if object_match:
        return json.loads(object_match.group(1))
    array_match = re.search(r"(\[.*\])", clean_text, flags=re.DOTALL)
    if array_match:
        return json.loads(array_match.group(1))
    raise ValueError("无法解析 JSON")


def _request_with_retry(method, url, max_retries=3, **kwargs):
    kwargs.setdefault("timeout", 30)
    last_error = None
    for attempt in range(1, max_retries + 1):
        try:
            if method == "post":
                return requests.post(url, **kwargs)
            return requests.get(url, **kwargs)
        except (requests.exceptions.ConnectionError, requests.exceptions.Timeout) as e:
            last_error = e
            if attempt < max_retries:
                wait = min(2 ** attempt, 10)
                print(f"    ⚠️ {url.split('/')[-1]} 第 {attempt} 次请求失败: {e}, {wait}s 后重试...")
                time.sleep(wait)
            else:
                raise last_error


def get_json_file_content(category_guid):
    if not category_guid:
        raise ValueError("category_guid 不能为空")
    signed_url_response = requests.get(
        BASE_URL + SIGNED_URL_ROUTE,
        headers=get_headers_with_ak(),
        params={"categoryGuid": category_guid},
        timeout=30,
    )
    signed_url_json = signed_url_response.json()
    signed_url = (signed_url_json.get("data") or {}).get("signedUrl")
    if not signed_url:
        raise Exception(f"获取 JSON 文件 signedUrl 失败: {signed_url_json}")
    file_response = requests.get(signed_url, timeout=60)
    if file_response.status_code != 200:
        raise Exception(f"下载 JSON 文件失败: status={file_response.status_code}, text={file_response.text[:300]}")
    text = file_response.text.strip()
    if not text:
        raise ValueError("JSON 文件内容为空")
    return safe_json_loads(text)


def load_prompt_text(prompt_file_guid, default_prompt):
    if not prompt_file_guid:
        return default_prompt
    try:
        signed_url_response = requests.get(
            BASE_URL + SIGNED_URL_ROUTE,
            headers=get_headers_with_ak(),
            params={"categoryGuid": prompt_file_guid},
            timeout=30,
        )
        signed_url = (signed_url_response.json().get("data") or {}).get("signedUrl")
        if not signed_url:
            return default_prompt
        return requests.get(signed_url, timeout=20).text
    except Exception as e:
        print(f"⚠️ Prompt 文件读取失败，使用默认 prompt: {e}")
        return default_prompt


def get_last_week_info():
    today = datetime.now()
    last_monday = today - timedelta(days=today.weekday() + 7)
    week_dates = [last_monday + timedelta(days=i) for i in range(7)]
    return {
        "start_date": week_dates[0].strftime("%Y-%m-%d"),
        "end_date": week_dates[-1].strftime("%Y-%m-%d"),
        "date_list": [d.strftime("%Y-%m-%d") for d in week_dates],
        "week_number": week_dates[0].isocalendar()[1],
        "week_key": f"{week_dates[0].strftime('%Y')}#W{week_dates[0].isocalendar()[1]:02d}",
    }


def build_rollup_note_title(week_info, dept_name):
    return f"{week_info['week_key']} {dept_name}周报"


def build_intermediate_json_file(name_prefix, target_date_str, json_content, suffix=""):
    tmp_dir = tempfile.gettempdir()
    unique_suffix = uuid.uuid4().hex[:8]
    name_suffix = f"_{suffix}" if suffix else ""
    file_name = f"{name_prefix}_{target_date_str.replace('-', '')}{name_suffix}_{unique_suffix}.json"
    file_path = os.path.join(tmp_dir, file_name)
    with open(file_path, "w", encoding="utf-8") as output_fp:
        json.dump(json_content, output_fp, ensure_ascii=False, indent=2)
    return file_path


def build_intermediate_markdown_file(name_prefix, target_date_str, markdown_content):
    tmp_dir = tempfile.gettempdir()
    unique_suffix = uuid.uuid4().hex[:8]
    file_name = f"{name_prefix}_{target_date_str.replace('-', '')}_{unique_suffix}.md"
    file_path = os.path.join(tmp_dir, file_name)
    with open(file_path, "w", encoding="utf-8") as output_fp:
        output_fp.write(markdown_content)
    return file_path


def cleanup_temp_files(file_paths, project_name=""):
    if not file_paths:
        return
    if bool(config.get("keep_temp_files", False)):
        print(f"[Cleanup][{project_name}] keep_temp_files=True，保留临时文件: {file_paths}")
        return
    for file_path in file_paths:
        try:
            if file_path and os.path.exists(file_path):
                os.remove(file_path)
                prefix = f"[Cleanup][{project_name}]" if project_name else "[Cleanup]"
                print(f"{prefix} 🧹 已删除临时文件: {file_path}")
        except Exception as e:
            prefix = f"[Cleanup][{project_name}]" if project_name else "[Cleanup]"
            print(f"{prefix} ⚠️ 删除临时文件失败: {file_path}, error={e}")


# =============================================================================
# OpenAI SDK 模型调用
# =============================================================================
def call_llm(messages, max_tokens=None, temperature=None, stream=True, max_retries=None, print_stream=None):
    max_tokens = max_tokens or LLM_MAX_TOKENS
    temperature = LLM_TEMPERATURE if temperature is None else temperature
    max_retries = max_retries or LLM_MAX_RETRIES
    should_print_stream = LLM_PRINT_STREAM if print_stream is None else bool(print_stream)

    last_error = None
    for attempt in range(1, max_retries + 1):
        try:
            print(f"    🔄 [LLM 尝试 {attempt}/{max_retries}] 调用 {LLM_MODEL} ...")
            if not stream:
                response = openai_client.chat.completions.create(
                    model=LLM_MODEL,
                    messages=messages,
                    max_tokens=max_tokens,
                    temperature=temperature,
                    stream=False,
                )
                return (response.choices[0].message.content or "").strip()

            response = openai_client.chat.completions.create(
                model=LLM_MODEL,
                messages=messages,
                max_tokens=max_tokens,
                temperature=temperature,
                stream=True,
            )
            chunks = []
            if should_print_stream:
                print("    🟢 开始流式输出：")
            for chunk in response:
                if not chunk.choices:
                    continue
                delta = chunk.choices[0].delta
                content = getattr(delta, "content", None)
                if content is None:
                    continue
                chunks.append(content)
                if should_print_stream:
                    print(content, end="", flush=True)
            if should_print_stream:
                print("\n    ✅ 流式输出结束")
            else:
                print("    ✅ LLM 流式调用完成")
            return "".join(chunks).strip()
        except Exception as e:
            last_error = e
            if attempt < max_retries:
                wait_time = min(2 ** (attempt - 1), 30)
                print(f"    ⚠️ LLM 调用失败: {e}. {wait_time}s 后重试...")
                time.sleep(wait_time)
            else:
                print(f"    ❌ LLM 连续 {max_retries} 次失败: {e}")
                raise last_error


# =============================================================================
# 部门配置表读取
# =============================================================================
REQUIRED_DEPT_COLUMNS = [
    "dept_id",
    "dept_name",
    "parent_dept_id",
    "dept_level",
    "state_project_guid",
    "state_parent_guid",
    "weekly_target_project_guid",
    "weekly_target_parent_guid",
    "owner_name",
    "owner_guid",
]


def normalize_dept_row(row):
    normalized = {}
    for k, v in (row or {}).items():
        key = str(k).strip()
        value = "" if v is None else str(v).strip()
        normalized[key] = value
    if "state_project_id" in normalized and "state_project_guid" not in normalized:
        normalized["state_project_guid"] = normalized["state_project_id"]
    for col in REQUIRED_DEPT_COLUMNS:
        normalized.setdefault(col, "")
    return normalized


def load_dept_config_from_xlsx(path):
    try:
        import openpyxl
    except Exception as e:
        raise ImportError("读取 xlsx 需要 openpyxl，请确认平台环境已安装") from e
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    headers = [str(cell.value).strip() if cell.value is not None else "" for cell in ws[1]]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(x is not None and str(x).strip() for x in row):
            continue
        raw = {headers[i]: row[i] for i in range(min(len(headers), len(row))) if headers[i]}
        rows.append(normalize_dept_row(raw))
    return rows


def load_dept_config_from_csv(path):
    rows = []
    with open(path, "r", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for row in reader:
            rows.append(normalize_dept_row(row))
    return rows


def download_config_file_by_guid(file_guid):
    signed_url_response = requests.get(
        BASE_URL + SIGNED_URL_ROUTE,
        headers=get_headers_with_ak(),
        params={"categoryGuid": file_guid},
        timeout=30,
    )
    signed_url = (signed_url_response.json().get("data") or {}).get("signedUrl")
    if not signed_url:
        raise Exception(f"获取部门配置文件 signedUrl 失败: {signed_url_response.text[:300]}")
    resp = requests.get(signed_url, timeout=60)
    if resp.status_code != 200:
        raise Exception(f"下载部门配置文件失败: {resp.status_code}, {resp.text[:300]}")
    suffix = ".xlsx"
    content_disposition = resp.headers.get("Content-Disposition", "")
    if ".csv" in content_disposition.lower():
        suffix = ".csv"
    tmp_path = os.path.join(tempfile.gettempdir(), f"dept_rollup_config_{uuid.uuid4().hex[:8]}{suffix}")
    with open(tmp_path, "wb") as f:
        f.write(resp.content)
    return tmp_path


def load_dept_config():
    if isinstance(config.get("departments"), list) and config.get("departments"):
        return [normalize_dept_row(x) for x in config["departments"]]

    candidate_path = ""
    if "dept_config_file" in globals():
        try:
            candidate_path = dept_config_file.path
        except Exception:
            candidate_path = ""
    if not candidate_path:
        candidate_path = config.get("dept_config_path", "")
    if not candidate_path and config.get("dept_config_file_guid"):
        candidate_path = download_config_file_by_guid(config["dept_config_file_guid"])
    if not candidate_path:
        raise ValueError("未找到部门配置表。请提供 config.departments、dept_config_file.path、dept_config_path 或 dept_config_file_guid")

    suffix = Path(candidate_path).suffix.lower()
    if suffix in (".xlsx", ".xlsm"):
        rows = load_dept_config_from_xlsx(candidate_path)
    elif suffix == ".csv":
        rows = load_dept_config_from_csv(candidate_path)
    else:
        raise ValueError(f"不支持的部门配置文件类型: {suffix}")
    if not rows:
        raise ValueError("部门配置表为空")
    return rows


def parse_dept_level(level):
    text = str(level or "").strip()
    match = re.search(r"\d+", text)
    if match:
        return int(match.group(0))
    return 0


def build_org_maps(dept_rows):
    dept_by_id = {}
    children_map = defaultdict(list)
    for row in dept_rows:
        dept_id = row.get("dept_id")
        if dept_id:
            dept_by_id[dept_id] = row
    for row in dept_rows:
        parent_id = row.get("parent_dept_id")
        if parent_id:
            children_map[parent_id].append(row)
    return dept_by_id, children_map


def get_leaf_descendants(target_dept_id, dept_by_id, children_map):
    children = children_map.get(target_dept_id, [])
    if not children:
        return [dept_by_id[target_dept_id]] if target_dept_id in dept_by_id else []
    leaves = []
    for child in children:
        leaves.extend(get_leaf_descendants(child["dept_id"], dept_by_id, children_map))
    return leaves


def get_parent_depts(dept_rows, children_map):
    parent_ids = set(children_map.keys())
    return [row for row in dept_rows if row.get("dept_id") in parent_ids]


# =============================================================================
# 原子池文件读取
# =============================================================================
def _get_tree_node_title(node):
    for key in ("dataTitle", "title", "name", "fileName", "filename"):
        value = node.get(key)
        if isinstance(value, str) and value.strip():
            return value.strip()
    return ""


def _get_tree_node_guid(node):
    for key in ("categoryGuid", "dataGuid", "guid", "fileGuid", "id"):
        value = node.get(key)
        if value:
            return value
    return ""


def _is_json_atomic_file_node(node):
    node_type = node.get("dataType", node.get("type"))
    try:
        is_file = int(node_type) == 5
    except Exception:
        is_file = str(node_type) == "5"
    title = _get_tree_node_title(node)
    return is_file and title.lower().endswith(".json")


def _infer_date_from_title(title, date_list):
    title = title or ""
    for date_str in date_list:
        variants = [date_str, date_str.replace("-", "/"), date_str.replace("-", "."), date_str.replace("-", "")]
        if any(v in title for v in variants):
            return date_str
    return ""


def find_weekly_atomic_files(user_guid, project_guid, folder_guid, date_list):
    response = requests.post(
        url=BASE_URL + DOC_TREE_ROUTE,
        headers=get_headers_with_ak(user_guid=user_guid),
        json={"projectGuid": project_guid, "parentGuid": folder_guid},
        timeout=60,
    )
    response_json = response.json()
    note_list = response_json.get("data") or []
    matched = []
    for note in note_list:
        if not _is_json_atomic_file_node(note):
            continue
        title = _get_tree_node_title(note)
        guid = _get_tree_node_guid(note)
        if not guid:
            continue
        matched.append({
            "date": _infer_date_from_title(title, date_list),
            "categoryGuid": guid,
            "dataTitle": title,
            "node_type": note.get("dataType", note.get("type")),
        })
    return matched


def get_item_project_name(item):
    direct = (item.get("project_name") or "").strip()
    if direct:
        return direct
    project = item.get("project", {}) or {}
    return (project.get("name") or "未分类项目").strip()


def normalize_section(section):
    section = (section or "progress").strip().lower()
    if section in ("progress", "main_progress", "done", "completed"):
        return "progress"
    if section in (
        "issue", "issues", "risk", "risks", "help", "issue_help", "issue_and_help",
        "issues_help", "risk_help", "support", "difficulty", "blocked", "blocker",
        "need_help", "needhelp",
    ):
        return "issues_support"
    if section in ("next", "next_focus", "next_plan", "plan", "todo", "future", "nextkeyfocus", "key_focus"):
        return "next_plan"
    return section


def strip_mentions_for_rollup(text):
    if not text:
        return ""
    text = re.sub(r"\[@[^\]]+\]\(mention:[^)]+\)", "", text)
    text = re.sub(r"@[^\s，。；、:：]+", "", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def clean_item_for_rollup(item):
    new_item = deepcopy(item)
    new_item["member"] = {"uid": "", "id": "", "label": "", "mention_md": ""}
    for block in new_item.get("content", []) or []:
        block["text"] = strip_mentions_for_rollup(block.get("text", ""))
        block["mentions"] = []
    return new_item


def load_leaf_weekly_atomic_pool(leaf_dept, week_info):
    project_name = leaf_dept.get("dept_name", "UnknownLeaf")
    state_project_guid = leaf_dept.get("state_project_guid")
    state_parent_guid = leaf_dept.get("state_parent_guid")
    if not state_project_guid or not state_parent_guid:
        print(f"    [Skip][{project_name}] 缺少 state_project_guid/state_parent_guid")
        return {"items": [], "metadata": {}}

    user_guid = leaf_dept.get("owner_guid") or USER_GUID
    files = find_weekly_atomic_files(user_guid, state_project_guid, state_parent_guid, week_info["date_list"])
    all_items = []
    source_urls = OrderedDict()
    for file_info in files:
        try:
            atomic_pool = get_json_file_content(file_info["categoryGuid"])
        except Exception as e:
            print(f"    [Skip][{project_name}] 原子池读取失败: {file_info.get('dataTitle')} error={e}")
            continue
        items = atomic_pool.get("items", []) or []
        meta = atomic_pool.get("meta", {}) or {}
        note_date = meta.get("date") or file_info.get("date")
        if note_date and note_date not in week_info["date_list"]:
            continue
        if not note_date:
            item_dates = sorted({x.get("date") for x in items if x.get("date")})
            week_item_dates = [d for d in item_dates if d in week_info["date_list"]]
            if week_item_dates:
                note_date = week_item_dates[0]
            else:
                continue
        source_report_url = meta.get("source_url") or meta.get("markdown_note_url") or meta.get("json_note_url") or f"{BASE_URL}/workspace/{file_info['categoryGuid']}"
        source_urls[note_date] = source_report_url
        for raw_item in items:
            item = clean_item_for_rollup(raw_item)
            if not item.get("date"):
                item["date"] = note_date
            if item.get("date") not in week_info["date_list"]:
                continue
            item["leaf_dept_id"] = leaf_dept.get("dept_id", "")
            item["leaf_dept_name"] = leaf_dept.get("dept_name", "")
            item["project_name"] = get_item_project_name(item)
            item["section"] = normalize_section(item.get("section"))
            all_items.append(item)
    return {
        "metadata": {
            "dept_id": leaf_dept.get("dept_id", ""),
            "dept_name": leaf_dept.get("dept_name", ""),
            "source_urls": dict(source_urls),
            "total_items": len(all_items),
        },
        "items": all_items,
    }


# =============================================================================
# content depth 树结构
# =============================================================================
def restore_content_tree(content_blocks):
    roots = []
    stack = []
    for block in content_blocks or []:
        text = (block.get("text") or "").strip()
        if not text:
            continue
        depth = block.get("depth")
        try:
            depth = int(depth)
        except Exception:
            depth = 0
        node = {"text": text, "block_type": block.get("block_type", "paragraph"), "depth": depth, "children": []}
        while stack and depth <= stack[-1][0]:
            stack.pop()
        if stack:
            stack[-1][1]["children"].append(node)
        else:
            roots.append(node)
        stack.append((depth, node))
    return roots


def tree_to_markdown(nodes, level=0):
    lines = []
    indent = "  " * level
    for node in nodes or []:
        lines.append(f"{indent}- {node.get('text', '')}")
        if node.get("children"):
            lines.extend(tree_to_markdown(node["children"], level + 1))
    return lines


def atomic_item_to_rollup_entry(item):
    source = item.get("source", {}) or {}
    return {
        "item_id": item.get("item_id") or str(uuid.uuid4()),
        "date": item.get("date", ""),
        "dept_name": item.get("dept_name", ""),
        "leaf_dept_id": item.get("leaf_dept_id", ""),
        "leaf_dept_name": item.get("leaf_dept_name", ""),
        "project_name": get_item_project_name(item),
        "section": normalize_section(item.get("section")),
        "content_tree": restore_content_tree(item.get("content", []) or []),
        "source": {
            "note_guid": source.get("note_guid", ""),
            "note_title": source.get("note_title", ""),
            "source_url": source.get("source_url", ""),
        },
    }


def build_rollup_timeline_state(rollup_atomic_pool, target_dept, week_info):
    items = rollup_atomic_pool.get("items", []) or []
    platform_map = OrderedDict()
    empty_platform_as_dept = bool(get_rollup_config(target_dept, "rollup_empty_platform_as_dept", True))
    for item in items:
        if item.get("platforms"):
            platforms = item.get("platforms") or []
        else:
            fallback = item.get("leaf_dept_name") if empty_platform_as_dept else "部门工作"
            platforms = [fallback or "部门工作"]
        entry = atomic_item_to_rollup_entry(item)
        project_name = entry["project_name"]
        section = entry["section"]
        date = entry["date"]
        leaf_key = (entry.get("leaf_dept_id", ""), entry.get("leaf_dept_name", "未知子部门"))
        for platform in platforms:
            platform = (platform or "部门工作").strip()
            if platform not in platform_map:
                platform_map[platform] = OrderedDict()
            if project_name not in platform_map[platform]:
                platform_map[platform][project_name] = {
                    "project_name": project_name,
                    "sections": {"progress": [], "issues_support": [], "next_plan": []},
                }
            platform_map[platform][project_name]["sections"].setdefault(section, [])
            platform_map[platform][project_name]["sections"][section].append({
                "leaf_dept_id": leaf_key[0],
                "leaf_dept_name": leaf_key[1],
                "date": date,
                "items": [entry],
            })
    platforms = []
    for platform_name, project_map in platform_map.items():
        platforms.append({"platform": platform_name, "projects": list(project_map.values())})
    return {
        "target_dept": {"dept_id": target_dept.get("dept_id", ""), "dept_name": target_dept.get("dept_name", "")},
        "date_range": {"start": week_info["start_date"], "end": week_info["end_date"], "week": week_info["week_key"]},
        "platforms": platforms,
    }


def compact_rollup_timeline_for_llm(batch_state, max_text_len_per_tree=2000):
    compact = {"platform": batch_state.get("platform"), "projects": []}
    for project in batch_state.get("projects", []) or []:
        compact_project = {"project_name": project.get("project_name", "未分类项目"), "sections": {}}
        for section_name, entries in (project.get("sections", {}) or {}).items():
            compact_entries = []
            for entry in entries or []:
                day_items = []
                for item in entry.get("items", []) or []:
                    tree_md = "\n".join(tree_to_markdown(item.get("content_tree", [])))
                    if len(tree_md) > max_text_len_per_tree:
                        tree_md = tree_md[:max_text_len_per_tree] + "..."
                    day_items.append({"item_id": item.get("item_id", ""), "content_tree_markdown": tree_md})
                if day_items:
                    compact_entries.append({
                        "leaf_dept_id": entry.get("leaf_dept_id", ""),
                        "leaf_dept_name": entry.get("leaf_dept_name", "未知子部门"),
                        "date": entry.get("date"),
                        "items": day_items,
                    })
            compact_project["sections"][section_name] = compact_entries
        compact["projects"].append(compact_project)
    return compact


def estimate_text_size(obj):
    return len(json.dumps(obj, ensure_ascii=False))


def split_large_project_by_leaf_dept(platform, project, max_batch_chars=12000):
    project_name = project.get("project_name", "未分类项目")
    sections = project.get("sections", {}) or {}
    leaf_map = OrderedDict()
    for section_name, entries in sections.items():
        for entry in entries or []:
            leaf_dept_id = entry.get("leaf_dept_id", "")
            leaf_dept_name = entry.get("leaf_dept_name", "未知子部门")
            key = (leaf_dept_id, leaf_dept_name)
            if key not in leaf_map:
                leaf_map[key] = {"leaf_dept_id": leaf_dept_id, "leaf_dept_name": leaf_dept_name, "sections": {"progress": [], "issues_support": [], "next_plan": []}}
            leaf_map[key]["sections"].setdefault(section_name, [])
            leaf_map[key]["sections"][section_name].append(entry)
    batches = []
    current_leaf_entries = []
    current_size = 0
    for leaf_entry in leaf_map.values():
        leaf_size = estimate_text_size(leaf_entry)
        if current_size + leaf_size > max_batch_chars and current_leaf_entries:
            batches.append(build_project_leaf_batch(platform, project_name, current_leaf_entries))
            current_leaf_entries = []
            current_size = 0
        current_leaf_entries.append(leaf_entry)
        current_size += leaf_size
    if current_leaf_entries:
        batches.append(build_project_leaf_batch(platform, project_name, current_leaf_entries))
    return batches


def build_project_leaf_batch(platform, project_name, leaf_entries):
    project = {"project_name": project_name, "sections": {"progress": [], "issues_support": [], "next_plan": []}}
    for leaf in leaf_entries:
        for section_name, entries in leaf.get("sections", {}).items():
            for entry in entries:
                project["sections"].setdefault(section_name, [])
                project["sections"][section_name].append(entry)
    return {"platform": platform, "projects": [project]}


def split_rollup_batches(rollup_timeline_state, max_projects_per_batch=5, max_batch_chars=12000):
    batches = []
    for platform_data in rollup_timeline_state.get("platforms", []) or []:
        platform = platform_data.get("platform", "部门工作")
        projects = platform_data.get("projects", [])
        current_projects = []
        current_size = 0
        for project in projects:
            project_size = estimate_text_size(project)
            if project_size > max_batch_chars:
                if current_projects:
                    batches.append({"platform": platform, "projects": current_projects})
                    current_projects = []
                    current_size = 0
                batches.extend(split_large_project_by_leaf_dept(platform, project, max_batch_chars=max_batch_chars))
                continue
            if len(current_projects) >= max_projects_per_batch or current_size + project_size > max_batch_chars:
                if current_projects:
                    batches.append({"platform": platform, "projects": current_projects})
                current_projects = []
                current_size = 0
            current_projects.append(project)
            current_size += project_size
        if current_projects:
            batches.append({"platform": platform, "projects": current_projects})
    return batches



def get_rollup_prompt_guid(target_dept, key):
    """
    Prompt GUID 读取优先级：
    1. target_dept[key]：部门级特殊 prompt
    2. config["rollup_global"][key]：逐层上报全局 prompt
    3. config[key]：顶层兼容字段

    推荐第一版直接把 prompt guid 写在 config 顶层或 rollup_global 中。
    """
    if target_dept and target_dept.get(key):
        return target_dept.get(key)

    rollup_global = config.get("rollup_global", {}) or {}
    if rollup_global.get(key):
        return rollup_global.get(key)

    return config.get(key, "")


# =============================================================================
# Rollup Prompts
# =============================================================================
def get_default_rollup_card_prompt():
    return """你是飞书卡片摘要助手。

你的任务是：将输入的父部门逐层上报周报压缩成适合飞书消息卡片展示的简洁正文。

# 输出目标
输出一段 200~350 字以内的卡片摘要，重点让读者快速知道：
1. 本周最重要的进展是什么
2. 当前有哪些明确困难或所需帮助
3. 下一步计划是什么

# 输出要求
1. 不要输出 Markdown 标题语法，例如 #、##、###。
2. 可以使用加粗强调关键词。
3. 使用项目符号 `•` 组织内容。
4. 不要输出代码块。
5. 不要输出 JSON。
6. 不要输出日期范围和源日报链接。
7. 不要展开所有项目细节，只保留最高优先级信息。
8. 不要编造输入中不存在的事实。
9. 如果无明确困难，写"暂无明确阻塞性问题"。
10. 如果无明确下一步计划，写"暂无明确下一步计划"。
11. 如输入中存在关键数量或明确枚举成果，优先保留数量，例如"完成 8 个算法验证"，不要写成"完成多个算法"。

# 推荐格式
**本周重点**
• ...
• ...

**困难/帮助**
• ...

**下一步**
• ...

# 输入周报
{{markdown_content}}
"""


def generate_rollup_card_content(target_dept, long_markdown, week_info):
    dept_name = target_dept.get("dept_name", "")
    prompt_file_guid = get_rollup_prompt_guid(target_dept, "rollup_card_prompt_file_guid")
    prompt_text = load_prompt_text(prompt_file_guid, get_default_rollup_card_prompt())

    def fallback_format_content(content, max_len=20000):
        content = re.sub(
            r"^###\s+(.+?)\s*$",
            lambda m: f"**{m.group(1).strip()}**",
            content,
            flags=re.MULTILINE,
        )
        if len(content) > max_len:
            return content[:max_len] + "\n\n......\n[系统提示：AI 生成失败，此为自动截断的格式化预览]"
        return content

    start_date = week_info["start_date"]
    end_date = week_info["end_date"]
    summary_prefix = f"**本周摘要 | {start_date} 至 {end_date}**\n\n"
    meta_header = f"时间范围：{start_date} 至 {end_date} | 第{week_info['week_number']}周"

    card_input_limit = int(get_rollup_config(target_dept, "rollup_card_input_max_chars", 0))
    if card_input_limit and card_input_limit > 0:
        card_source_markdown = long_markdown[:card_input_limit]
    else:
        card_source_markdown = long_markdown

    card_input_markdown = f"{meta_header}\n\n{card_source_markdown}"
    user_content = prompt_text.replace("{{markdown_content}}", card_input_markdown)

    try:
        llm_result = call_llm(
            messages=[
                {"role": "system", "content": "你是内容整理助手，请综合父部门周报生成适合飞书卡片展示的精炼正文，不要输出代码块。"},
                {"role": "user", "content": user_content},
            ],
            max_tokens=int(get_rollup_config(target_dept, "rollup_card_max_tokens", 1024)),
            temperature=float(get_rollup_config(target_dept, "rollup_card_temperature", 0.3)),
            stream=True,
            max_retries=int(get_rollup_config(target_dept, "llm_max_retries", LLM_MAX_RETRIES)),
        )
        return summary_prefix + strip_markdown_wrapper(llm_result)
    except Exception as e:
        print(f"⚠️ [Card][{dept_name}] AI 生成失败，使用 fallback: {e}")
        return summary_prefix + fallback_format_content(card_input_markdown, max_len=20000)


def get_default_rollup_trend_prompt():
    return """你是父部门逐层上报趋势分析助手。

你会收到某一个 platform 下若干 project 的叶子部门原子事实。输入已经删除所有个人 mention 和个人信息。

你的任务是：基于输入事实，生成父部门周报使用的结构化 JSON。

# 必须输出合法 JSON，不要输出代码块，不要输出解释文字

# 输出结构
{
  "platform": "",
  "core_progress": [
    {
      "project_name": "",
      "subtopic": "",
      "involved_departments": [""],
      "summary": "",
      "evidence_dates": ["YYYY-MM-DD"],
      "source_item_ids": [""]
    }
  ],
  "main_progress": [],
  "issues_support": [],
  "next_plan": []
}

# 规则
1. platform 必须来自输入 platform。
2. project_name 必须来自输入 project_name，不允许根据正文改写。
3. subtopic 可以从 content_tree_markdown 中抽取，但不能替代 project_name。
4. involved_departments 必须来自 leaf_dept_name。
5. 不允许输出个人姓名、个人 mention、普通 @姓名。
6. 必须保留数量、完成度、枚举对象。例如 AT2、AT3、AT4、M0C、MT2、M01、M02、SNT、SNM 不得压缩为“多个 bank”。
7. issues_support 只输出明确困难、风险、阻塞、依赖、所需支持。
8. next_plan 只输出明确后续计划。
9. 不允许新增事实。
10. main_progress 尽量覆盖每个 project 的主要进展；core_progress 只选本批次最关键事项。

输入 JSON：
{{batch_json}}
"""


def get_default_rollup_final_prompt():
    return """你是父部门逐层上报周报整理助手。

你会收到结构化周报草稿，内容来自该父部门下所有叶子部门的 atomic facts，并且已经删除个人 mention。

你的任务是生成父部门周报正文。

# 最高优先级原则
1. 只基于输入内容整理，不允许新增事实。
2. 父部门周报面向管理层，不展示个人 mention。
3. 严禁输出 [@姓名](mention:uid:id)。
4. 严禁输出普通 @姓名。
5. 不输出子部门周报链接。
6. 必须保留涉及部门名称。
7. 必须保留 platform、project、数量、完成度、枚举对象。
8. 不要把多个独立项目合并成“相关工作”“多个项目”等模糊表述。

# 输出结构
必须严格输出以下 4 个章节：

### 🎉 本周关键进展

### ✅ 本周主要进展

### ❗ 困难及所需帮助

### 🙌 下一步计划

# 推荐层级
- **{Platform}**
    - **{Project}**（涉及部门：{DeptA、DeptB}）
        - **{Subtopic}**
            - 具体事项

如果没有明确 Subtopic，可直接在 Project 下列事项。

# 输入内容
{{markdown_content}}
"""


def get_default_rollup_validation_prompt():
    return """你是父部门逐层上报覆盖性校验助手。

你的任务是对照“结构化草稿”和“最终父部门周报”，检查最终周报是否遗漏 platform / project / involved_departments / 数量 / 困难 / 下一步计划。

# 输出要求
只输出合法 JSON。

# 输出结构
{
  "pass": true,
  "missing_items": [
    {
      "section": "本周主要进展 / 困难及所需帮助 / 下一步计划 / 其他",
      "platform": "",
      "project_name": "",
      "missing_fact": "",
      "suggested_insert_position": ""
    }
  ],
  "wrong_or_suspicious_items": [
    {
      "section": "",
      "issue": ""
    }
  ]
}

# 校验重点
1. 不允许出现个人 mention 或普通 @姓名。
2. 结构化草稿中明确的 Project 不能丢失。
3. involved_departments 不能丢失。
4. 明确数量、完成度、枚举对象不能被压缩成“多个/若干/相关”。
5. 明确困难/支持事项必须进入“困难及所需帮助”。
6. 明确下一步计划必须进入“下一步计划”。

# 结构化草稿
{{structured_markdown}}

# 最终周报
{{final_markdown}}
"""


def get_default_rollup_repair_prompt():
    return """你是父部门逐层上报周报修复助手。

你会收到：
1. 结构化草稿
2. 当前最终父部门周报
3. 覆盖性校验结果 JSON

请在不大规模重写的前提下，补回遗漏内容。

# 要求
1. 只补充结构化草稿中明确存在的事实。
2. 严禁输出个人 mention。
3. 严禁输出普通 @姓名。
4. 不输出日期范围、周数、源链接等 header。
5. 保留 platform / project / involved_departments / 数量 / 枚举对象。
6. 只输出修复后的 Markdown 正文。

# 结构化草稿
{{structured_markdown}}

# 当前最终周报
{{final_markdown}}

# 校验结果 JSON
{{validation_json}}
"""


# =============================================================================
# Rollup Trend / Final / Validation / Repair
# =============================================================================
def build_rollup_trend_prompt(batch_state, target_dept):
    prompt_template = load_prompt_text(
        get_rollup_prompt_guid(target_dept, "rollup_trend_prompt_file_guid"),
        get_default_rollup_trend_prompt(),
    )
    compact = compact_rollup_timeline_for_llm(
        batch_state,
        max_text_len_per_tree=int(get_rollup_config(target_dept, "rollup_tree_max_chars", 2000)),
    )
    return prompt_template.replace("{{batch_json}}", json.dumps(compact, ensure_ascii=False, indent=2))


def fallback_rollup_analyze_batch(batch_state):
    platform = batch_state.get("platform", "部门工作")
    result = {"platform": platform, "core_progress": [], "main_progress": [], "issues_support": [], "next_plan": []}
    for project in batch_state.get("projects", []) or []:
        project_name = project.get("project_name", "未分类项目")
        for section_name, target_key in [("progress", "main_progress"), ("issues_support", "issues_support"), ("next_plan", "next_plan")]:
            grouped = []
            involved_depts = set()
            dates = set()
            ids = []
            for entry in project.get("sections", {}).get(section_name, []) or []:
                involved_depts.add(entry.get("leaf_dept_name", "未知子部门"))
                if entry.get("date"):
                    dates.add(entry["date"])
                for item in entry.get("items", []) or []:
                    ids.append(item.get("item_id", ""))
                    tree_md = "\n".join(tree_to_markdown(item.get("content_tree", [])))
                    if tree_md:
                        grouped.append(tree_md)
            if grouped:
                summary = "；".join(grouped[:8])
                if len(summary) > 800:
                    summary = summary[:800] + "..."
                result[target_key].append({
                    "project_name": project_name,
                    "subtopic": project_name,
                    "involved_departments": sorted(involved_depts),
                    "summary": summary,
                    "evidence_dates": sorted(dates),
                    "source_item_ids": [x for x in ids if x],
                })
    result["core_progress"] = result["main_progress"][:3]
    return result


def analyze_rollup_batch(batch_idx, total_batches, batch_state, target_dept):
    dept_name = target_dept.get("dept_name", "UnknownDept")
    platform = batch_state.get("platform", "部门工作")
    print(f"    [Rollup Trend Batch {batch_idx}/{total_batches}][{safe_log_text(dept_name)}][{safe_log_text(platform)}] 开始分析，项目数: {len(batch_state.get('projects', []))}")
    messages = [
        {"role": "system", "content": "你是父部门逐层上报趋势分析助手，只输出合法 JSON。"},
        {"role": "user", "content": build_rollup_trend_prompt(batch_state, target_dept)},
    ]
    try:
        llm_result = call_llm(
            messages=messages,
            max_tokens=int(get_rollup_config(target_dept, "rollup_trend_max_tokens", LLM_MAX_TOKENS)),
            temperature=float(get_rollup_config(target_dept, "rollup_trend_temperature", 0.1)),
            stream=True,
            max_retries=int(get_rollup_config(target_dept, "llm_max_retries", LLM_MAX_RETRIES)),
        )
        parsed = safe_json_loads(llm_result)
        parsed.setdefault("platform", platform)
        return batch_idx, parsed
    except Exception as e:
        print(f"    [Rollup Trend Batch {batch_idx}/{total_batches}][{platform}] ⚠️ 分析失败，使用 fallback: {e}")
        return batch_idx, fallback_rollup_analyze_batch(batch_state)


def analyze_rollup_trends_in_parallel(rollup_timeline_state, target_dept):
    max_projects = int(get_rollup_config(target_dept, "rollup_projects_per_batch", 5))
    max_chars = int(get_rollup_config(target_dept, "rollup_max_batch_chars", 12000))
    max_parallel = min(int(get_rollup_config(target_dept, "batch_number", 10)), 50)
    batches = split_rollup_batches(rollup_timeline_state, max_projects_per_batch=max_projects, max_batch_chars=max_chars)
    if not batches:
        return []
    total = len(batches)
    actual_parallel = min(max_parallel, total)
    print(f"[Step 3][{target_dept.get('dept_name')}] 开始父部门 rollup 分块趋势分析，共 {total} 批，并行数: {actual_parallel}")
    results = {}
    with ThreadPoolExecutor(max_workers=actual_parallel) as executor:
        future_to_idx = {executor.submit(analyze_rollup_batch, idx, total, batch, target_dept): idx for idx, batch in enumerate(batches, 1)}
        for future in as_completed(future_to_idx):
            idx = future_to_idx[future]
            try:
                finished_idx, result = future.result()
                results[finished_idx] = result
            except Exception as e:
                print(f"    [Rollup Trend Batch {idx}/{total}] ❌ 失败: {e}")
                raise
    return [results[i] for i in range(1, total + 1)]


def normalize_rollup_item(item):
    return {
        "project_name": (item.get("project_name") or "未分类项目").strip(),
        "subtopic": (item.get("subtopic") or item.get("project_name") or "未分类事项").strip(),
        "involved_departments": item.get("involved_departments", []) or [],
        "summary": (item.get("summary") or "").strip(),
        "evidence_dates": item.get("evidence_dates", []) or [],
        "source_item_ids": item.get("source_item_ids", []) or [],
    }


def merge_rollup_trend_results(batch_results):
    platform_map = OrderedDict()
    section_keys = ["core_progress", "main_progress", "issues_support", "next_plan"]
    for result in batch_results or []:
        platform = result.get("platform") or "部门工作"
        if platform not in platform_map:
            platform_map[platform] = {key: [] for key in section_keys}
        for key in section_keys:
            for item in result.get(key, []) or []:
                norm = normalize_rollup_item(item)
                if norm["summary"]:
                    platform_map[platform][key].append(norm)
    return platform_map


def dedupe_rollup_items(items):
    seen = set()
    out = []
    for item in items or []:
        key = (item.get("project_name"), item.get("subtopic"), tuple(item.get("involved_departments", [])), item.get("summary"))
        if key in seen:
            continue
        seen.add(key)
        out.append(item)
    return out


def render_rollup_platform_section(platform_map, section_key, empty_text="本周暂无明确内容。"):
    parts = []
    any_content = False
    for platform, sections in platform_map.items():
        items = dedupe_rollup_items(sections.get(section_key, []) or [])
        if not items:
            continue
        any_content = True
        parts.append(f"- **{platform}**")
        grouped_by_project = OrderedDict()
        for item in items:
            grouped_by_project.setdefault(item["project_name"], [])
            grouped_by_project[item["project_name"]].append(item)
        for project_name, project_items in grouped_by_project.items():
            involved_all = []
            for item in project_items:
                involved_all.extend(item.get("involved_departments", []))
            involved_unique = sorted({x for x in involved_all if x})
            dept_suffix = f"（涉及部门：{'、'.join(involved_unique)}）" if involved_unique else ""
            parts.append(f"    - **{project_name}**{dept_suffix}")
            grouped_by_subtopic = OrderedDict()
            for item in project_items:
                subtopic = item.get("subtopic") or project_name
                grouped_by_subtopic.setdefault(subtopic, [])
                grouped_by_subtopic[subtopic].append(item)
            for subtopic, sub_items in grouped_by_subtopic.items():
                if subtopic and subtopic != project_name:
                    parts.append(f"        - **{subtopic}**")
                    leaf_indent = "            "
                else:
                    leaf_indent = "        "
                for item in sub_items:
                    parts.append(f"{leaf_indent}- {item['summary']}")
        parts.append("")
    if not any_content:
        return empty_text
    return "\n".join(parts).strip()


def build_rollup_structured_body(platform_map):
    parts = []
    parts.append("### 🎉 本周关键进展")
    parts.append(render_rollup_platform_section(platform_map, "core_progress", "本周暂无核心产出。"))
    parts.append("")
    parts.append("### ✅ 本周主要进展")
    parts.append(render_rollup_platform_section(platform_map, "main_progress", "本周暂无明确主要进展。"))
    parts.append("")
    parts.append("### ❗ 困难及所需帮助")
    parts.append(render_rollup_platform_section(platform_map, "issues_support", "本周无明确阻塞性问题或外部协助事项。"))
    parts.append("")
    parts.append("### 🙌 下一步计划")
    parts.append(render_rollup_platform_section(platform_map, "next_plan", "本周无明确下一步计划。"))
    parts.append("")
    return "\n".join(parts).strip()


def build_rollup_header(target_dept, leaf_depts, week_info):
    leaf_names = [x.get("dept_name", "") for x in leaf_depts if x.get("dept_name")]
    if len(leaf_names) <= 8:
        scope_text = "、".join(leaf_names)
    else:
        scope_text = f"{'、'.join(leaf_names[:8])} 等 {len(leaf_names)} 个叶子部门"
    parts = [
        f"**日期范围：** {week_info['start_date']} 至 {week_info['end_date']} | **周数：** 第 {week_info['week_number']} 周",
        "",
        f"**汇总范围：** {scope_text}",
        "",
        "---",
    ]
    return "\n".join(parts)


def generate_rollup_final_body(structured_body, target_dept):
    prompt_template = load_prompt_text(
        get_rollup_prompt_guid(target_dept, "rollup_final_prompt_file_guid"),
        get_default_rollup_final_prompt(),
    )
    prompt = prompt_template.replace("{{markdown_content}}", structured_body)
    try:
        result = call_llm(
            messages=[
                {"role": "system", "content": "你是父部门逐层上报周报整理助手，请只输出 Markdown 正文，不要输出个人 mention。"},
                {"role": "user", "content": prompt},
            ],
            max_tokens=int(get_rollup_config(target_dept, "rollup_final_max_tokens", LLM_MAX_TOKENS)),
            temperature=float(get_rollup_config(target_dept, "rollup_final_temperature", 0.2)),
            stream=True,
            max_retries=int(get_rollup_config(target_dept, "llm_max_retries", LLM_MAX_RETRIES)),
        )
        return strip_markdown_wrapper(result) or structured_body
    except Exception as e:
        print(f"    ⚠️ 父部门最终整理失败，使用结构化草稿: {e}")
        return structured_body


def validate_rollup_coverage(structured_body, final_body, target_dept):
    prompt_template = load_prompt_text(
        get_rollup_prompt_guid(target_dept, "rollup_validation_prompt_file_guid"),
        get_default_rollup_validation_prompt(),
    )
    prompt = prompt_template.replace("{{structured_markdown}}", structured_body).replace("{{final_markdown}}", final_body)
    try:
        result = call_llm(
            messages=[
                {"role": "system", "content": "你是父部门逐层上报覆盖性校验助手，只输出合法 JSON。"},
                {"role": "user", "content": prompt},
            ],
            max_tokens=int(get_rollup_config(target_dept, "rollup_validation_max_tokens", 2048)),
            temperature=float(get_rollup_config(target_dept, "rollup_validation_temperature", 0.0)),
            stream=True,
            max_retries=int(get_rollup_config(target_dept, "llm_max_retries", LLM_MAX_RETRIES)),
        )
        parsed = safe_json_loads(result)
        parsed.setdefault("pass", True)
        parsed.setdefault("missing_items", [])
        parsed.setdefault("wrong_or_suspicious_items", [])
        return parsed
    except Exception as e:
        print(f"    ⚠️ 父部门覆盖校验失败，跳过校验: {e}")
        return {"pass": True, "missing_items": [], "wrong_or_suspicious_items": [], "skipped": True}


def repair_rollup_body(structured_body, final_body, validation_result, target_dept):
    prompt_template = load_prompt_text(
        get_rollup_prompt_guid(target_dept, "rollup_repair_prompt_file_guid"),
        get_default_rollup_repair_prompt(),
    )
    prompt = prompt_template.replace("{{structured_markdown}}", structured_body).replace("{{final_markdown}}", final_body).replace("{{validation_json}}", json.dumps(validation_result, ensure_ascii=False, indent=2))
    try:
        result = call_llm(
            messages=[
                {"role": "system", "content": "你是父部门逐层上报周报修复助手，请只输出修复后的 Markdown 正文。"},
                {"role": "user", "content": prompt},
            ],
            max_tokens=int(get_rollup_config(target_dept, "rollup_repair_max_tokens", LLM_MAX_TOKENS)),
            temperature=float(get_rollup_config(target_dept, "rollup_repair_temperature", 0.1)),
            stream=True,
            max_retries=int(get_rollup_config(target_dept, "llm_max_retries", LLM_MAX_RETRIES)),
        )
        return strip_markdown_wrapper(result) or final_body
    except Exception as e:
        print(f"    ⚠️ 父部门修复失败，保留原正文: {e}")
        return final_body


def generate_checked_rollup_body(structured_body, target_dept):
    final_body = generate_rollup_final_body(structured_body, target_dept)
    if not get_rollup_config(target_dept, "enable_rollup_coverage_check", True):
        return final_body, {"pass": True, "missing_items": [], "wrong_or_suspicious_items": [], "skipped": True}
    validation = validate_rollup_coverage(structured_body, final_body, target_dept)
    missing = validation.get("missing_items", []) or []
    suspicious = validation.get("wrong_or_suspicious_items", []) or []
    if validation.get("pass", True) and not missing and not suspicious:
        print("    ✅ 父部门覆盖性校验通过")
        return final_body, validation
    print(f"    ⚠️ 父部门校验发现遗漏/可疑项：missing={len(missing)}, suspicious={len(suspicious)}，开始修复")
    repaired = repair_rollup_body(structured_body, final_body, validation, target_dept)
    if get_rollup_config(target_dept, "enable_rollup_second_validation", False):
        second = validate_rollup_coverage(structured_body, repaired, target_dept)
        return repaired, second
    return repaired, validation


# =============================================================================
# 文档写入
# =============================================================================
def _convert_special_nodes(content):
    content = re.sub(
        r"\[([^\]]+)\]\(mentionUrl:[^:]+:[^:]+:([^)]+)\)",
        lambda m: f'<a data-node-type="mentionUrl" data-url="{m.group(2)}">{m.group(1)}</a>',
        content,
    )
    content = re.sub(
        r"\[([^\]]+)\]\((https?://[^)\s]+)\)",
        lambda m: f'<a href="{m.group(2)}">{m.group(1)}</a>',
        content,
    )
    return content


def insert_markdown_to_note(user_guid, note_guid, markdown_content, max_retries=3):
    clean_content = strip_markdown_wrapper(markdown_content)
    html_content = _convert_special_nodes(clean_content)
    response = _request_with_retry(
        "post",
        BASE_URL + MD_INSERT_ROUTE,
        max_retries=max_retries,
        headers=get_headers_with_ak(user_guid=user_guid),
        json={"note_guid": note_guid, "markdown_content": html_content, "mode": "w", "location": 1},
        timeout=60,
    )
    if response.status_code != 200:
        raise Exception(f"写入笔记失败: {response.text}")
    return response.json()


def create_note_api(content, title, project_guid, parent_guid, tags, creator_guid=None):
    creator_guid = creator_guid or USER_GUID
    headers = get_headers_with_ak()
    headers["X-User-GUID"] = creator_guid
    if not project_guid:
        raise ValueError("weekly_target_project_guid 不能为空")
    response = _request_with_retry(
        "post",
        BASE_URL + WORKSPACE_SAVE_ROUTE,
        max_retries=3,
        headers=headers,
        json={"project_guid": project_guid, "parent_guid": parent_guid, "target": {"name": title, "type": 1, "tags": tags}, "creator_guid": creator_guid},
        timeout=60,
    )
    response_json = response.json()
    if response.status_code != 200 or not response_json.get("data"):
        raise Exception(f"创建笔记 API 返回错误: {response_json}")
    doc_id = response_json.get("data", {}).get("guid")
    if doc_id and content:
        try:
            insert_markdown_to_note(creator_guid, doc_id, content, max_retries=5)
        except Exception as e:
            print(f"    ⚠️ 笔记已创建(doc_id={doc_id})但内容写入失败: {e}")
            time.sleep(5)
            insert_markdown_to_note(creator_guid, doc_id, content, max_retries=5)
    return doc_id


def create_rollup_weekly_note(content, target_dept, week_info):
    dept_name = target_dept.get("dept_name", "")
    target_project_guid = target_dept.get("weekly_target_project_guid")
    target_parent_guid = target_dept.get("weekly_target_parent_guid", "0")
    creator_guid = target_dept.get("owner_guid") or USER_GUID
    print(f"[Step 5][{dept_name}] 正在创建逐层上报周报...")
    title = build_rollup_note_title(week_info, dept_name)
    doc_id = create_note_api(
        content=content,
        title=title,
        project_guid=target_project_guid,
        parent_guid=target_parent_guid,
        tags=["周报", "AI"],
        creator_guid=creator_guid,
    )
    note_url = f"{BASE_URL}/workspace/{doc_id}" if doc_id else ""
    print(f"[Step 5][{dept_name}] ✅ 逐层上报周报创建完成: {note_url}")
    return note_url, title



# =============================================================================
# Webhook 群消息发送：父部门逐层上报简单卡片
# =============================================================================
def parse_webhook_urls(value):
    """
    webhook_url 支持：
    - 单个 URL
    - 多个 URL：用 ; , ， ； 换行 分隔
    - list[str]
    """
    if not value:
        return []

    if isinstance(value, list):
        return [str(x).strip() for x in value if str(x).strip()]

    text = str(value).strip()
    if not text:
        return []

    parts = re.split(r"[;,；，\n]+", text)
    return [x.strip() for x in parts if x.strip()]


def build_rollup_message_text(note_title, note_url):
    return f"【{note_title}】已生成，请点击查看。\\n<a href='{note_url}'>点击查看详情</a>"


def build_rollup_simple_card(title, content, note_url):
    return {
        "schema": "2.0",
        "header": {
            "padding": "12px 8px 12px 8px",
            "template": "blue",
            "title": {
                "content": title,
                "tag": "plain_text"
            }
        },
        "body": {
            "vertical_spacing": "12px",
            "elements": [
                {
                    "tag": "markdown",
                    "content": content,
                    "margin": "0px",
                    "text_size": "normal"
                },
                {
                    "tag": "column_set",
                    "flex_mode": "stretch",
                    "horizontal_spacing": "8px",
                    "margin": "8px 0px 0px 0px",
                    "columns": [
                        {
                            "tag": "column",
                            "width": "auto",
                            "elements": [
                                {
                                    "tag": "button",
                                    "type": "primary_filled",
                                    "width": "fill",
                                    "margin": "4px 0px 4px 0px",
                                    "text": {
                                        "tag": "plain_text",
                                        "content": "查看完整周报"
                                    },
                                    "behaviors": [
                                        {
                                            "type": "open_url",
                                            "default_url": note_url
                                        }
                                    ]
                                }
                            ]
                        }
                    ]
                }
            ]
        }
    }


def send_webhook(webhook_url, card):
    response = requests.post(
        url=webhook_url,
        headers={"Content-Type": "application/json"},
        json={"msg_type": "interactive", "card": card},
        timeout=30
    )
    try:
        return response.json()
    except Exception:
        return {"status_code": response.status_code, "text": response.text[:300]}


def send_message_api(receiver_guids, title, content, sender_guid="", interactive_content=None):
    payload = {
        "template_id": MESSAGE_TEMPLATE_ID,
        "receiver_guid": receiver_guids,
        "content": content,
        "org_guid": ORG_GUID,
        "title": title,
        "platform_type": PLATFORM_TYPE,
    }
    if interactive_content is not None:
        payload["interactive_content"] = json.dumps(interactive_content, ensure_ascii=False)
    return requests.post(
        url=BASE_URL + MESSAGE_SEND_ROUTE,
        headers=get_headers_with_ak(user_guid=sender_guid),
        json=payload,
        timeout=30,
    )


def send_rollup_webhook_messages(target_dept, note_title, note_url, card_content=None):
    dept_name = target_dept.get("dept_name", "")
    webhook_urls = parse_webhook_urls(target_dept.get("webhook_url", ""))
    owner_guid = (target_dept.get("owner_guid") or "").strip()

    content = card_content or build_rollup_message_text(note_title, note_url)
    card = build_rollup_simple_card(note_title, content, note_url)

    # --- 群消息：Webhook ---
    if webhook_urls:
        for idx, webhook_url in enumerate(webhook_urls, 1):
            try:
                print(f"[Step 6][{dept_name}] 正在发送父部门周报卡片 Webhook {idx}/{len(webhook_urls)}")
                result = send_webhook(webhook_url, card)
                if result.get("code") == 0 or result.get("StatusCode") == 0:
                    print(f"  -> ✅ Webhook 发送成功: {webhook_url[:30]}...")
                else:
                    print(f"  -> ⚠️ Webhook 返回异常: {result}")
            except Exception as e:
                print(f"  -> ❌ Webhook 发送失败: {e}")
    else:
        print(f"[Step 6][{dept_name}] 未配置 webhook_url，跳过群消息发送")

    # --- 个人消息：发送给 owner_guid ---
    if owner_guid:
        try:
            print(f"[Step 6][{dept_name}] 正在发送个人消息给 owner ({owner_guid[:8]}...)")
            text_content = build_rollup_message_text(note_title, note_url)
            sender_guid = owner_guid
            response = send_message_api(
                receiver_guids=[owner_guid],
                title=note_title,
                content=text_content,
                sender_guid=sender_guid,
                interactive_content=card,
            )
            if response.status_code == 200 and response.json().get("data"):
                print(f"  -> ✅ 个人消息发送成功")
            else:
                print(f"  -> ❌ 个人消息发送失败: {response.text[:200]}")
        except Exception as e:
            print(f"  -> ❌ 个人消息发送异常: {e}")
    else:
        print(f"[Step 6][{dept_name}] 未配置 owner_guid，跳过个人消息发送")



# =============================================================================
# 主流程：从叶子 atomic JSON 直接生成目标父部门逐层上报周报
# =============================================================================
def build_rollup_atomic_pool_for_target(target_dept, leaf_depts, week_info):
    all_items = []
    leaf_stats = []
    print(f"[Step 1][{target_dept.get('dept_name')}] 目标父部门下叶子部门数: {len(leaf_depts)}")
    for leaf in leaf_depts:
        leaf_pool = load_leaf_weekly_atomic_pool(leaf, week_info)
        items = leaf_pool.get("items", []) or []
        all_items.extend(items)
        leaf_stats.append({"dept_id": leaf.get("dept_id", ""), "dept_name": leaf.get("dept_name", ""), "item_count": len(items)})
        print(f"    - {safe_log_text(leaf.get('dept_name'))}: {len(items)} items")
    return {
        "metadata": {
            "target_dept_id": target_dept.get("dept_id", ""),
            "target_dept_name": target_dept.get("dept_name", ""),
            "range_dates": week_info["date_list"],
            "week_number": week_info["week_number"],
            "week_key": week_info["week_key"],
            "leaf_depts": [{"dept_id": x.get("dept_id", ""), "dept_name": x.get("dept_name", "")} for x in leaf_depts],
            "leaf_stats": leaf_stats,
            "total_items": len(all_items),
        },
        "items": all_items,
    }


def run_rollup_for_target_dept(target_dept, dept_by_id, children_map, week_info):
    dept_name = target_dept.get("dept_name", "")
    temp_files = []
    try:
        leaf_depts = get_leaf_descendants(target_dept["dept_id"], dept_by_id, children_map)
        if len(leaf_depts) == 1 and leaf_depts[0].get("dept_id") == target_dept.get("dept_id"):
            print(f"[Skip][{dept_name}] 是叶子节点，逐层上报应用不处理叶子节点")
            return
        rollup_atomic_pool = build_rollup_atomic_pool_for_target(target_dept, leaf_depts, week_info)
        if not rollup_atomic_pool.get("items"):
            print(f"[Skip][{dept_name}] 未读取到任何叶子 atomic items")
            return
        raw_json_path = build_intermediate_json_file(f"rollup_{target_dept.get('dept_id')}", f"{week_info['start_date']}_to_{week_info['end_date']}", rollup_atomic_pool, suffix="atomic_raw")
        temp_files.append(raw_json_path)
        rollup_timeline_state = build_rollup_timeline_state(rollup_atomic_pool, target_dept, week_info)
        timeline_path = build_intermediate_json_file(f"rollup_{target_dept.get('dept_id')}", f"{week_info['start_date']}_to_{week_info['end_date']}", rollup_timeline_state, suffix="timeline")
        temp_files.append(timeline_path)
        batch_results = analyze_rollup_trends_in_parallel(rollup_timeline_state, target_dept)
        batch_path = build_intermediate_json_file(f"rollup_{target_dept.get('dept_id')}", f"{week_info['start_date']}_to_{week_info['end_date']}", {"batches": batch_results}, suffix="trend_batches")
        temp_files.append(batch_path)
        platform_map = merge_rollup_trend_results(batch_results)
        structured_body = build_rollup_structured_body(platform_map)
        structured_path = build_intermediate_markdown_file(f"rollup_{target_dept.get('dept_id')}", f"{week_info['start_date']}_to_{week_info['end_date']}_structured", structured_body)
        temp_files.append(structured_path)
        final_body, validation = generate_checked_rollup_body(structured_body, target_dept)
        header = build_rollup_header(target_dept, leaf_depts, week_info)
        final_markdown = header + "\n\n" + final_body.strip()
        final_path = build_intermediate_markdown_file(f"rollup_{target_dept.get('dept_id')}", f"{week_info['start_date']}_to_{week_info['end_date']}_final", final_markdown)
        temp_files.append(final_path)
        validation_path = build_intermediate_json_file(f"rollup_{target_dept.get('dept_id')}", f"{week_info['start_date']}_to_{week_info['end_date']}", validation, suffix="validation")
        temp_files.append(validation_path)
        note_url, title = create_rollup_weekly_note(final_markdown, target_dept, week_info)
        card_content = generate_rollup_card_content(target_dept, final_markdown, week_info)
        send_rollup_webhook_messages(target_dept, title, note_url or "", card_content=card_content)
        print(f"✅ {dept_name} 逐层上报周报完成: {note_url}")
    except Exception as e:
        print(f"❌ {dept_name} 逐层上报流程中断: {e}")
        traceback.print_exc()
    finally:
        cleanup_temp_files(temp_files, project_name=dept_name)


# =============================================================================
# 顶层执行：云端环境无需 main
# =============================================================================
print("=" * 60)
print("开始执行逐层上报周报工作流（从叶子 atomic JSON 动态汇总，无 weekly_state）")
print("=" * 60)

print("Prompt 配置：")
print(f"  rollup_trend_prompt_file_guid: {get_rollup_prompt_guid({}, 'rollup_trend_prompt_file_guid') or '使用代码兜底 prompt'}")
print(f"  rollup_final_prompt_file_guid: {get_rollup_prompt_guid({}, 'rollup_final_prompt_file_guid') or '使用代码兜底 prompt'}")
print(f"  rollup_validation_prompt_file_guid: {get_rollup_prompt_guid({}, 'rollup_validation_prompt_file_guid') or '使用代码兜底 prompt'}")
print(f"  rollup_repair_prompt_file_guid: {get_rollup_prompt_guid({}, 'rollup_repair_prompt_file_guid') or '使用代码兜底 prompt'}")
print(f"  rollup_card_prompt_file_guid: {get_rollup_prompt_guid({}, 'rollup_card_prompt_file_guid') or '使用代码兜底 prompt'}")


week_info = get_last_week_info()
dept_rows = load_dept_config()
dept_by_id, children_map = build_org_maps(dept_rows)

target_dept_ids = normalize_to_list(config.get("rollup_target_dept_ids", []))
if target_dept_ids:
    target_depts = []
    for dept_id in target_dept_ids:
        if dept_id not in dept_by_id:
            print(f"⚠️ rollup_target_dept_id 不存在: {dept_id}")
            continue
        target_depts.append(dept_by_id[dept_id])
else:
    target_depts = get_parent_depts(dept_rows, children_map)
    target_depts = sorted(target_depts, key=lambda x: parse_dept_level(x.get("dept_level")), reverse=True)

print(f"目标周期: {week_info['start_date']} ~ {week_info['end_date']}")
print(f"目标父部门数: {len(target_depts)}")
for target_dept in target_depts:
    run_rollup_for_target_dept(target_dept, dept_by_id, children_map, week_info)

print("\n" + "=" * 60)
print("全部逐层上报周报任务执行完毕")
print("=" * 60)
